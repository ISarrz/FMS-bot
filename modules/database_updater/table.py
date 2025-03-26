import re

import openpyxl


class Table:
    def __init__(self, sheet):
        self.sheet = sheet
        self.matrix = self.generate_matrix()
        self.width = self.sheet.max_column
        self.height = self.sheet.max_row

    def generate_matrix(self):
        matrix = [[''] * self.sheet.max_column for row in range(self.sheet.max_row)]

        for row in range(self.sheet.max_row):
            for col in range(self.sheet.max_column):
                matrix[row][col] = str(self.sheet[row + 1][col].value)

        for rng in self.sheet.merged_cells.ranges:
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col - 1, rng.max_col):
                    matrix[row - 1][col] = str(rng.start_cell.value)

        return matrix

    def find_cell_by_value(self, value):
        for row in range(self.height):
            for col in range(self.width):
                if self.matrix[row][col] == value:
                    return row, col

        return None

    def find_cell_by_regular_pattern(self, pattern):
        for row in range(self.height):
            for col in range(self.width):

                if re.fullmatch(pattern, self.matrix[row][col], re.IGNORECASE):
                    return row, col

        return None
